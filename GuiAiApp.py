import tkinter as tk
from tkinter import messagebox, filedialog
import openai
import PyPDF2
import openpyxl

openai.api_key = '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'

# Function chat_gpt that takes prompt as input, uses ChatGPT API to analyze prompt, and returns API response as output
def dataAnalyzer(prompt):
    # Call OpenAI chat completion API to generate a response from prompt
    response = openai.chat.completions.create(
        model="gpt-3.5-turbo", # GPT model
        messages=[
            # System role with description of being a "AI Data Analyzer Assistant" to interpret user's input data
            {"role": "system", "content": "As an AI Data Analyzer Assistant, your task is to provide a comprehensive analysis of the input text, capturing all key information and important context. Your analysis should include both statistical and distribution analyses to offer a holistic understanding of the data. For statistical analysis, provide metrics such as Mean, Median, Mode, Range, Variance, Standard Deviation, Quartiles, Percentiles, and Skewness."},
            # Role of user, takes user input as the prompt
            {"role": "user", "content": prompt}
            ]
    )
    # Return content of the generated response
    return response.choices[0].message.content.strip()

def contentSummarizer(prompt):
    # Call OpenAI chat completion API to generate a response from prompt
    response = openai.chat.completions.create(
        model="gpt-3.5-turbo", # GPT model
        messages=[
            # System role with description of being a "AI Data Analyzer Assistant" to interpret user's input data
            {"role": "system", "content": "You are an AI Content Summarizer Assistant. Your task is to provide a concise summary of the input text while capturing all key information and important context. Include who, what, where, and why as needed"},
            # Role of user, takes user input as the prompt
            {"role": "user", "content": prompt}
            ]
    )
    # Return content of the generated response
    return response.choices[0].message.content.strip()

# Function to handle button click event for opening files
def open_file():
    file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
    if file_path:
        try:
            if file_path.endswith('.txt'):
                with open(file_path, 'r') as txt_file:
                    content = txt_file.read()
            elif file_path.endswith('.pdf'):
                pdf_file = open(file_path, 'rb')
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                page_count = len(pdf_reader.pages)
                content = ""
                for page_num in range(page_count):
                    page = pdf_reader.pages[page_num]
                    content += page.extract_text()
                pdf_file.close()
            elif file_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file_path)
                content = ""
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        content += ' '.join(str(cell) for cell in row) + '\n'
        except Exception as e:
            messagebox.showerror("Error", str(e))
        else:
            input_text.delete("1.0", tk.END)
            input_text.insert(tk.END, content)

# Function to perform analysis based on user choice
def perform_analysis(choice, prompt):
    if choice == '1':
        # Generate data analysis using dataAnalyzer function
        return dataAnalyzer(prompt)
    elif choice == '2':
        # Generate content summary using contentSummarizer function
        return contentSummarizer(prompt)


# Function to handle button click event
def analyze_content():
    # Get user input from text entry
    user_prompt = input_text.get("1.0", tk.END).strip()
    # Get user choice from radio buttons
    choice = choice_var.get()
    # Perform analysis based on user input and choice
    result = perform_analysis(choice, user_prompt)
    # Display result in message box
    messagebox.showinfo("Analysis Result", result)

# Create main Tkinter window
root = tk.Tk()
root.title("Data Analysis and Content Summary Tool")

# Create text entry for user input with word wrapping
input_frame = tk.Frame(root)
input_frame.pack(pady=10)
input_text = tk.Text(input_frame, height=15, width=75, wrap="word")
input_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Create scrollbar for the text entry widget
scrollbar = tk.Scrollbar(input_frame, command=input_text.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
input_text.config(yscrollcommand=scrollbar.set)

# Create button to open file
open_file_button = tk.Button(root, text="Open File", command=open_file)
open_file_button.pack()

# Create radio buttons for user choice
choice_var = tk.StringVar()
choice_var.set("1")
tk.Radiobutton(root, text="Data Analysis", variable=choice_var, value="1").pack()
tk.Radiobutton(root, text="Content Summary", variable=choice_var, value="2").pack()

# Create button to perform analysis
analyze_button = tk.Button(root, text="Analyze Content", command=analyze_content)
analyze_button.pack(pady=10)

# Run Tkinter event loop
root.mainloop()
